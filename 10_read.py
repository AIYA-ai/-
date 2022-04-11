import json
import openpyxl as op

with open('evaluate.tang.json',encoding='utf8') as fp:
    json_tang_data = json.load(fp)
with open('evaluate.song.json',encoding='utf8') as fp:
    json_song_data = json.load(fp)
print(json_tang_data)
print(json_song_data)
bg = op.load_workbook(r"zancun.xlsx")  # 应先将excel文件放入到工作目录下
sheet = bg["Sheet1"]  # “Sheet1”表示将数据写入到excel文件的sheet1下
a1 = []
a2 = []
a1.extend(json_tang_data.keys())
a2.extend(json_song_data.keys())
for i in range(1000):
    sheet.cell(i + 2, 1, a1[i])
    sheet.cell(i + 2, 2, json_tang_data[a1[i]])
    sheet.cell(i + 2, 3, a2[i])
    sheet.cell(i + 2, 4, json_song_data[a2[i]])
bg.save("zancun.xlsx")  # 对文件进行保存