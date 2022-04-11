# 该文件用于实现共现矩阵统计词共现
import json
import os
import jieba
import openpyxl as op

path = './json/'
files = os.listdir(path)
name_s = []
name_t = []
ss = []
record = []
rows = 0

with open('evaluate.out.song.json', 'r', encoding='utf8')as fp:  # 读取数据
    json_song_ev = json.load(fp)
json_song_ev = list(json_song_ev)
with open('evaluate.out.tang.json', 'r', encoding='utf8')as fp:  # 读取数据
    json_tang_ev = json.load(fp)
json_tang_ev = list(json_tang_ev)
sss = []

exclude = ['[', ']', ' ', ',', "'", '。', '，', '《', '》', '）', '（', '「', '」', '：', '', '？', '□', '、', '：', '〖',
           '〗', '？', '“', '”', '『', '』', '']  # 无效字符排除

s_matrix = [[0 for xi in range(1000)] for xi in range(1000)]
t_matrix = [[0 for xi in range(1000)] for xi in range(1000)]
for file in files:
    s_name = 'song'
    t_name = 'tang'
    if s_name in file:
        new_name = path + file
        with open(new_name, 'r', encoding='utf8')as fp:  # 读取数据
            json_data = json.load(fp)
        for i in range(len(json_data)):
            reply_number = json_data[i]['paragraphs']
            words2 = str(reply_number)
            words1 = jieba.lcut(words2)
            ss = []
            for chara in words1:
                if chara not in json_song_ev:
                    continue
                if chara in exclude:
                    continue
                if chara not in ss:
                    ss.append(chara)
                else:
                    continue
                if len(chara) > 2:
                    continue

            for ii in ss:
                for jj in ss:
                    if ii == jj:
                        continue
                    s_matrix[json_song_ev.index(ii)][json_song_ev.index(jj)] += 1

    elif t_name in file:
        new_name = path + file
        with open(new_name, 'r', encoding='utf8')as fp:  # 读取数据
            json_data = json.load(fp)
        for i in range(len(json_data)):
            reply_number = json_data[i]['paragraphs']
            words2 = str(reply_number)
            words1 = jieba.lcut(words2)
            ss = []
            for chara in words1:
                if chara not in json_tang_ev:
                    continue
                if chara in exclude:
                    continue
                if chara not in ss:
                    ss.append(chara)
                else:
                    continue
                if len(chara) > 2:
                    continue

            for ii in ss:
                for jj in ss:
                    if ii == jj:
                        continue
                    t_matrix[json_tang_ev.index(ii)][json_tang_ev.index(jj)] += 1
    rows += 1
    print('已运行至第' + str(rows) + '个文件')  # 输出第多少个文件
bg = op.load_workbook(r"t_matrix.xlsx")  # 应先将excel文件放入到工作目录下
sheet = bg["Sheet1"]  # “Sheet1”表示将数据写入到excel文件的sheet1下
for i in json_tang_ev:
    for j in json_tang_ev:
        sss.append([i, j, t_matrix[json_tang_ev.index(i)][json_tang_ev.index(j)]])
        sheet.cell(json_tang_ev.index(i)+2, json_tang_ev.index(j)+2,
                   t_matrix[json_tang_ev.index(i)][json_tang_ev.index(j)])
    sheet.cell(json_tang_ev.index(i) + 2, 1, i)
    sheet.cell(1, json_tang_ev.index(i) + 2, i)
data1 = sorted(sss, key=lambda x: x[2], reverse=True)  # 排序
bg.save("t_matrix.xlsx")  # 对文件进行保存
sss = []
bg = op.load_workbook(r"s_matrix.xlsx")  # 应先将excel文件放入到工作目录下
sheet = bg["Sheet1"]  # “Sheet1”表示将数据写入到excel文件的sheet1下
for i in json_song_ev:
    for j in json_song_ev:
        sss.append([i, j, s_matrix[json_song_ev.index(i)][json_song_ev.index(j)]])
        sheet.cell(json_song_ev.index(i) + 2, json_song_ev.index(j) + 2,
                   s_matrix[json_song_ev.index(i)][json_song_ev.index(j)])
    sheet.cell(json_song_ev.index(i) + 2, 1, i)
    sheet.cell(1, json_song_ev.index(i) + 2, i)
bg.save("s_matrix.xlsx")  # 对文件进行保存
data2 = sorted(sss, key=lambda x: x[2], reverse=True)  # 排序
filed = 'occurrence.tang.times.json'
file_w = open(filed, 'w')
json_str = json.dumps(data1)
file_w.write(json_str)
file_w.close()
filed = 'occurrence.song.times.json'
file_w = open(filed, 'w')
json_str = json.dumps(data2)
file_w.write(json_str)
file_w.close()
